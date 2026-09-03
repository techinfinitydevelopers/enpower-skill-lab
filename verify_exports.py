"""
Check the list exports.

  python verify_exports.py

For each registry entry: the right role gets a real workbook whose headers
match the list screen, every other role is refused, and a scoped export
contains only that user's own school. The last one matters most -- a School
Admin exporting another school's parents would be a data leak, and nothing
else in the suite looks for it.

Passwords it changes to sign in are restored from atexit.
"""

import atexit
import io
import os
import sys

import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'enpower_skill_lab.settings')
django.setup()

from django.conf import settings                          # noqa: E402

settings.ALLOWED_HOSTS = list(settings.ALLOWED_HOSTS) + ['testserver']

from django.contrib.auth import get_user_model            # noqa: E402

from enpower_skill_lab.exports import _registry            # noqa: E402
from verify_client import HttpsClient as Client            # noqa: E402

U = get_user_model()
PASS, FAIL = [], []
restore = {}
TEST_PASSWORD = 'ExportCheck!2026'

ALL_ROLES = ('SUPER_ADMIN', 'PROGRAM_COORDINATOR', 'SCHOOL_ADMIN',
             'THINKING_COACH', 'PARENT', 'STUDENT')


def _restore_passwords():
    if not restore:
        return
    for pk, password in list(restore.items()):
        U.objects.filter(pk=pk).update(password=password)
    print(f'\n  restored original passwords for {len(restore)} user(s)')
    restore.clear()


atexit.register(_restore_passwords)


def check(label, ok, detail=''):
    (PASS if ok else FAIL).append(label)
    print(f'  {"PASS" if ok else "FAIL"}  {label}{("  - " + detail) if detail else ""}')


def client_for(role):
    # For a scoped role, pick an account that actually has a school attached --
    # otherwise the scoping test silently skips, which is the one check here
    # that guards against a cross-school data leak.
    user = None
    if role == 'SCHOOL_ADMIN':
        from school_admin.models import SchoolAdmin
        # A school with no students would make the scoping test vacuous, so
        # prefer one that actually has some.
        base = SchoolAdmin.objects.select_related('user', 'school').filter(
            school__isnull=False, user__isnull=False, user__is_active=True)
        profile = (base.filter(school__students__isnull=False).distinct().first()
                   or base.first())
        user = profile.user if profile else None
    elif role == 'THINKING_COACH':
        from teacher.models import Teacher
        profile = (Teacher.objects.select_related('user', 'school')
                   .filter(school__isnull=False, user__isnull=False,
                           user__is_active=True).first())
        user = profile.user if profile else None
    user = user or U.objects.filter(role=role, is_active=True).first()
    if not user:
        return None, None
    restore.setdefault(user.pk, U.objects.get(pk=user.pk).password)
    user.set_password(TEST_PASSWORD)
    user.save(update_fields=['password'])
    c = Client()
    if not c.login(username=user.username, password=TEST_PASSWORD):
        return None, None
    return c, user


def workbook_from(response):
    from openpyxl import load_workbook
    raw = b''.join(response.streaming_content) if getattr(response, 'streaming', False) \
        else response.content
    return load_workbook(io.BytesIO(raw))


registry = _registry()
clients = {}
for role in ALL_ROLES:
    c, u = client_for(role)
    if c:
        clients[role] = (c, u)
    else:
        print(f'  ..    no {role} account on this database')

print(f'\n  {len(registry)} exports registered\n')

# ── each export works for its own role ──────────────────────────────────
print('THE RIGHT ROLE GETS A WORKBOOK')
for key, entry in registry.items():
    role = entry['roles'][0]
    if role not in clients:
        check(f'{key}: an account exists to test with', False, f'no {role}')
        continue
    c, _ = clients[role]
    r = c.get(f'/exports/{key}/')
    ok = r.status_code == 200 and 'spreadsheet' in r.headers.get('Content-Type', '')
    check(f'{key}: {role} gets an xlsx', ok,
          f'HTTP {r.status_code} {r.headers.get("Content-Type", "")[:34]}')
    if not ok:
        continue

    wb = workbook_from(r)
    ws = wb.active
    headers = [ws.cell(row=3, column=i + 1).value
               for i in range(len(entry['columns']))]
    expected = [h for h, _ in entry['columns']]
    check(f'{key}: headers match the list screen', headers == expected,
          f'{headers}')
    check(f'{key}: attachment filename is set',
          'attachment' in r.headers.get('Content-Disposition', ''))
    check(f'{key}: sheet is titled', bool(ws.title))

# ── every other role is refused ──────────────────────────────────────────
print('\nEVERY OTHER ROLE IS REFUSED')
for key, entry in registry.items():
    refused, allowed_through = [], []
    for role, (c, _) in clients.items():
        if role in entry['roles']:
            continue
        r = c.get(f'/exports/{key}/')
        if r.status_code == 200:
            allowed_through.append(f'{role} ({r.status_code})')
        else:
            refused.append(role)
    check(f'{key}: refused to {len(refused)} other role(s)',
          not allowed_through, '; '.join(allowed_through))

print('\nANONYMOUS')
anon = Client()
leaked = [k for k in registry if anon.get(f'/exports/{k}/').status_code == 200]
check('no export is readable without signing in', not leaked, '; '.join(leaked))

check('an unknown key is refused',
      clients['SUPER_ADMIN'][0].get('/exports/not-a-real-export/').status_code != 200
      if 'SUPER_ADMIN' in clients else False)

# ── scoping: a School Admin sees only their own school ──────────────────
print('\nSCOPING  (the check that matters)')
if 'SCHOOL_ADMIN' in clients:
    from school_admin.models import SchoolAdmin
    from student.models import Student

    c, user = clients['SCHOOL_ADMIN']
    profile = getattr(user, 'school_admin_profile', None)
    own = getattr(profile, 'school', None) if profile else None

    if own is None:
        print('  ..    that School Admin has no school attached; scoping not testable')
    else:
        r = c.get('/exports/my-students/')
        ws = workbook_from(r).active
        exported_ids = {ws.cell(row=i, column=1).value
                        for i in range(4, ws.max_row + 1)}
        exported_ids.discard(None)
        # An empty export writes a placeholder in the first cell; it is not an id.
        exported_ids.discard('No records to export')

        mine = set(Student.objects.filter(school=own)
                   .values_list('skill_lab_reg_id', flat=True))
        others = set(Student.objects.exclude(school=own)
                     .values_list('skill_lab_reg_id', flat=True))

        print(f'  school: {own.school_name}')
        print(f'  their students: {len(mine)}, other schools: {len(others)}')
        check('every exported student belongs to that school',
              not (exported_ids - mine),
              f'{len(exported_ids - mine)} from elsewhere')
        check('no student from another school is present',
              not (exported_ids & others),
              f'{len(exported_ids & others)} leaked')
        if others:
            check('the scoping is actually narrowing something',
                  len(exported_ids) < len(mine) + len(others))

# ── the button is on the page ───────────────────────────────────────────
# An export nobody can reach is not an export. Each list must actually
# render a link to its own key.
print('\nTHE BUTTON IS ON THE PAGE')
PAGES = {
    'schools':             ('SUPER_ADMIN', '/super-admin/schools/'),
    'students':            ('SUPER_ADMIN', '/super-admin/students/'),
    'teachers':            ('SUPER_ADMIN', '/super-admin/teachers/'),
    'parents':             ('SUPER_ADMIN', '/super-admin/parents/'),
    'coordinators':        ('SUPER_ADMIN', '/super-admin/coordinators/'),
    'school-admins':       ('SUPER_ADMIN', '/super-admin/school-admins/'),
    'coordinator-schools': ('PROGRAM_COORDINATOR', '/coordinator/school-list/'),
    'my-students':         ('SCHOOL_ADMIN', '/school-admin/students/'),
    'my-parents':          ('SCHOOL_ADMIN', '/school-admin/parents/'),
    'class-students':      ('THINKING_COACH', '/teacher/students/'),
}
for key, (role, url) in PAGES.items():
    if role not in clients:
        check(f'{key}: page reachable to test', False, f'no {role}')
        continue
    c, _ = clients[role]
    r = c.get(url, follow=True)
    if r.status_code != 200:
        check(f'{key}: list page loads', False, f'{url} HTTP {r.status_code}')
        continue
    body = r.content.decode(errors='ignore')
    check(f'{key}: Export link present on {url}',
          f'/exports/{key}/' in body)

_restore_passwords()

print('\n' + '=' * 62)
print(f'PASS {len(PASS)}   FAIL {len(FAIL)}')
for f in FAIL:
    print('  FAILED:', f)
sys.exit(1 if FAIL else 0)
