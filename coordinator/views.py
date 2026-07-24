from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth import logout, update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.db.models import Count
from django.shortcuts import get_object_or_404
from schools.models import School, Class
from teacher.models import Teacher
from student.models import Student
from parent.models import Parent
from accounts.models import User
from attendance.models import Timetable, TimetableSlot, GRADE_CHOICES, ACADEMIC_YEAR_CHOICES, DAY_CHOICES
from django.utils import timezone
from operator import attrgetter
import json


def _coordinator_schools(request):
    """Return queryset of schools assigned to the requesting SRM (Program Coordinator).
    Prefer ProgramCoordinator.schools_assigned, fall back to School.srm filter."""
    try:
        coordinator = request.user.program_coordinator
        schools = coordinator.schools_assigned.all()
        if schools.exists():
            return schools
    except Exception:
        pass
    return School.objects.filter(srm=request.user)


def is_coordinator(user):
    """Check if user is a program coordinator"""
    return user.is_authenticated and hasattr(user, 'role') and user.role == 'PROGRAM_COORDINATOR'


@login_required
@user_passes_test(is_coordinator)
def coordinator_dashboard(request):
    """Coordinator dashboard view"""
    # Get coordinator profile and assigned schools
    try:
        coordinator = request.user.program_coordinator
        assigned_schools = coordinator.schools_assigned.all()
    except Exception:
        assigned_schools = School.objects.none()

    assigned_school_ids = assigned_schools.values_list('id', flat=True)

    # Dynamic counts
    total_schools = assigned_schools.count()
    total_teachers = Teacher.objects.filter(school_id__in=assigned_school_ids).count()
    total_students = Student.objects.filter(school_id__in=assigned_school_ids).count()

    # Schools with teacher/student/class counts for the table
    schools_with_counts = assigned_schools.annotate(
        teacher_count=Count('teachers', distinct=True),
        student_count=Count('students', distinct=True),
        class_count=Count('classes', distinct=True),
    ).order_by('-created_at')[:5]

    # Add initials for avatar display
    for school in schools_with_counts:
        words = school.school_name.split()
        school.initials = ''.join([w[0].upper() for w in words[:2]]) if words else '??'

    # School-wise summary (same data, add total_count for display)
    school_summary = list(schools_with_counts)
    max_students = 1
    for school in school_summary:
        school.total_count = school.teacher_count + school.student_count + school.class_count
        if school.student_count > max_students:
            max_students = school.student_count

    # Recent Activities — merge recent teachers, students, classes from assigned schools
    recent_teachers = Teacher.objects.filter(school_id__in=assigned_school_ids).order_by('-created_at')[:5]
    recent_students = Student.objects.filter(school_id__in=assigned_school_ids).order_by('-created_at')[:5]
    recent_classes = Class.objects.filter(school_id__in=assigned_school_ids).order_by('-created_at')[:5]

    activities = []
    for t in recent_teachers:
        t.activity_type = 'teacher'
        t.title = f'New Teacher Added'
        t.description = f'{t.full_name} joined as {t.get_designation_display() if hasattr(t, "get_designation_display") else t.designation}'
        t.school_name = t.school.school_name if t.school else '—'
        activities.append(t)

    for s in recent_students:
        s.activity_type = 'student'
        s.title = f'New Student Enrolled'
        s.description = f'{s.first_name} {s.last_name} enrolled in Class {s.student_class or "—"}'
        s.school_name = s.school.school_name if s.school else '—'
        activities.append(s)

    for c in recent_classes:
        c.activity_type = 'class'
        c.title = f'New Class Created'
        c.description = f'{c.class_name} — {c.academic_year}'
        c.school_name = c.school.school_name if c.school else '—'
        activities.append(c)

    activities.sort(key=attrgetter('created_at'), reverse=True)
    recent_activities = activities[:5]

    context = {
        'total_schools': total_schools,
        'total_teachers': total_teachers,
        'total_students': total_students,
        'pending_alerts': 0,
        'assigned_schools': schools_with_counts,
        'school_summary': school_summary,
        'max_students': max_students,
        'recent_activities': recent_activities,
    }
    return render(request, 'coordinator/dashboard.html', context)


@login_required
@user_passes_test(is_coordinator)
def coordinator_profile(request):
    """Coordinator profile view with update functionality"""
    if request.method == 'POST':
        try:
            # Get or create program coordinator profile
            coordinator = request.user.program_coordinator

            # Handle profile photo upload
            if request.FILES.get('profile_photo'):
                coordinator.profile_photo = request.FILES['profile_photo']

            # Update coordinator fields
            full_name = request.POST.get('full_name', '').strip()
            if full_name:
                coordinator.full_name = full_name
                # Also update user's first and last name
                name_parts = full_name.split(' ', 1)
                request.user.first_name = name_parts[0]
                request.user.last_name = name_parts[1] if len(name_parts) > 1 else ''

            # Update phone numbers
            mobile_number = request.POST.get('mobile_number', '').strip()
            if mobile_number:
                coordinator.mobile_number = mobile_number

            alternate_number = request.POST.get('alternate_number', '').strip()
            if alternate_number:
                coordinator.alternate_number = alternate_number
            else:
                coordinator.alternate_number = None

            # Update gender
            gender = request.POST.get('gender', '').strip()
            if gender:
                coordinator.gender = gender

            # Update date of birth
            date_of_birth = request.POST.get('date_of_birth', '').strip()
            if date_of_birth:
                coordinator.date_of_birth = date_of_birth

            # Update email
            email = request.POST.get('email', '').strip()
            if email:
                coordinator.official_email = email
                request.user.email = email

            # Save changes
            coordinator.save()
            request.user.save()

            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('coordinator:coordinator_profile')

        except Exception as e:
            messages.error(request, f'Error updating profile: {str(e)}')
            return redirect('coordinator:coordinator_profile')

    return render(request, 'coordinator/profile.html')


@login_required
@user_passes_test(is_coordinator)
def coordinator_change_password(request):
    """Coordinator change password view"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)

            # Update last password change timestamp
            try:
                coordinator = request.user.program_coordinator
                coordinator.last_password_change = timezone.now()
                coordinator.save()
            except:
                pass  # Coordinator profile might not exist

            messages.success(request, 'Your password was successfully updated!')
            return redirect('coordinator:coordinator_profile')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'coordinator/change_password.html', {'form': form})


@login_required
def coordinator_logout(request):
    """Coordinator logout view"""
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


@login_required
@user_passes_test(is_coordinator)
def school_list(request):
    """School list view"""
    # Fetch only schools assigned to the requesting coordinator
    schools = _coordinator_schools(request).order_by('-created_at')

    context = {
        'schools': schools,
    }

    return render(request, 'coordinator/school-list.html', context)


# ============================================================
# TIMETABLE MANAGEMENT (PPT slides 12-13)
# ============================================================
def _day_label(code):
    """Map a day_of_week code (mon/tue/..) to its full name via DAY_CHOICES."""
    return dict(DAY_CHOICES).get(code, code)


def _fmt_time(t):
    """Format a TimeField as HH:MM, blank if None."""
    return t.strftime('%H:%M') if t else ''


@login_required
@user_passes_test(is_coordinator)
def timetable_list(request):
    """List timetables for the SRM's assigned schools, flattened to one row per slot."""
    assigned_schools = _coordinator_schools(request)
    school_ids = assigned_schools.values_list('id', flat=True)
    timetables = (
        Timetable.objects
        .filter(school_id__in=school_ids)
        .select_related('school', 'thinking_coach')
        .prefetch_related('slots')
        .order_by('-created_at')
    )

    rows = []
    for tt in timetables:
        teacher = ''
        if tt.thinking_coach:
            teacher = tt.thinking_coach.get_full_name() or tt.thinking_coach.username
        slots = list(tt.slots.all())
        if slots:
            for slot in slots:
                start = _fmt_time(slot.start_time)
                end = _fmt_time(slot.end_time)
                if start and end:
                    timings = f'{start} – {end}'
                elif start or end:
                    timings = start or end
                else:
                    timings = '—'
                rows.append({
                    'program': tt.program,
                    'grade': tt.grade,
                    'division': tt.division,
                    'day': _day_label(slot.day_of_week),
                    'timings': timings,
                    'start_date': tt.start_date,
                    'teacher': teacher,
                    'tt_id': tt.id,
                    'tt': tt,
                })
        else:
            rows.append({
                'program': tt.program,
                'grade': tt.grade,
                'division': tt.division,
                'day': '—',
                'timings': '—',
                'start_date': tt.start_date,
                'teacher': teacher,
                'tt_id': tt.id,
                'tt': tt,
            })

    context = {
        'rows': rows,
        'total_timetables': timetables.count(),
        'total_schools': assigned_schools.count(),
    }
    return render(request, 'coordinator/timetable-list.html', context)


def _save_slots(request, timetable):
    """Create TimetableSlot rows from posted slot_day/slot_period/slot_start/slot_end/slot_note."""
    days = request.POST.getlist('slot_day')
    periods = request.POST.getlist('slot_period')
    starts = request.POST.getlist('slot_start')
    ends = request.POST.getlist('slot_end')
    slot_notes = request.POST.getlist('slot_note')
    for i, day in enumerate(days):
        if not day:
            continue
        period = periods[i] if i < len(periods) and periods[i] else 1
        TimetableSlot.objects.create(
            timetable=timetable,
            day_of_week=day,
            period_number=period,
            start_time=(starts[i] if i < len(starts) and starts[i] else None),
            end_time=(ends[i] if i < len(ends) and ends[i] else None),
            note=(slot_notes[i] if i < len(slot_notes) else ''),
        )


@login_required
@user_passes_test(is_coordinator)
def timetable_detail(request, pk):
    """Read-only display of a full schedule. Scoped to coordinator's schools."""
    assigned_schools = _coordinator_schools(request)
    school_ids = assigned_schools.values_list('id', flat=True)
    timetable = (
        Timetable.objects
        .filter(id=pk, school_id__in=school_ids)
        .select_related('school', 'thinking_coach')
        .prefetch_related('slots')
        .first()
    )
    if not timetable:
        messages.error(request, 'Timetable not found or not in your assigned schools.')
        return redirect('coordinator:timetable_list')

    slot_rows = []
    for slot in timetable.slots.all():
        start = _fmt_time(slot.start_time)
        end = _fmt_time(slot.end_time)
        if start and end:
            timings = f'{start} – {end}'
        elif start or end:
            timings = start or end
        else:
            timings = '—'
        slot_rows.append({
            'day': _day_label(slot.day_of_week),
            'timings': timings,
            'note': slot.note,
        })

    context = {
        'timetable': timetable,
        'slot_rows': slot_rows,
    }
    return render(request, 'coordinator/timetable-detail.html', context)


@login_required
@user_passes_test(is_coordinator)
def timetable_edit(request, pk):
    """Edit an existing schedule. Reuses the create form (timetable-upload.html) in edit mode."""
    assigned_schools = _coordinator_schools(request)
    school_ids = assigned_schools.values_list('id', flat=True)
    timetable = (
        Timetable.objects
        .filter(id=pk, school_id__in=school_ids)
        .select_related('school', 'thinking_coach')
        .prefetch_related('slots')
        .first()
    )
    if not timetable:
        messages.error(request, 'Timetable not found or not in your assigned schools.')
        return redirect('coordinator:timetable_list')

    thinking_coaches = User.objects.filter(role='THINKING_COACH').order_by('first_name', 'username')

    if request.method == 'POST':
        school_id = request.POST.get('school', '').strip()
        coach_id = request.POST.get('thinking_coach', '').strip()
        grade = request.POST.get('grade', '').strip()
        division = request.POST.get('division', '').strip()
        academic_year = request.POST.get('academic_year', '').strip()
        program = request.POST.get('program', '').strip()
        start_date = request.POST.get('start_date', '').strip() or None
        end_date = request.POST.get('end_date', '').strip() or None
        notes = request.POST.get('notes', '').strip()
        schedule_file = request.FILES.get('schedule_file')

        if not school_id or not grade or not division:
            messages.error(request, 'School, grade and section are required.')
            return redirect('coordinator:timetable_edit', pk=pk)

        school = assigned_schools.filter(id=school_id).first()
        if not school:
            messages.error(request, 'Invalid school selection.')
            return redirect('coordinator:timetable_edit', pk=pk)

        coach = None
        if coach_id:
            coach = thinking_coaches.filter(id=coach_id).first()

        if not program:
            program = school.get_skill_program_display() if school.skill_program else ''

        try:
            timetable.school = school
            timetable.thinking_coach = coach
            timetable.grade = grade
            timetable.division = division
            timetable.academic_year = academic_year or '2025-2026'
            timetable.program = program
            timetable.start_date = start_date
            timetable.end_date = end_date
            if schedule_file:
                timetable.schedule_file = schedule_file
            timetable.notes = notes
            timetable.save()

            # Replace slots
            timetable.slots.all().delete()
            _save_slots(request, timetable)

            messages.success(request, 'Timetable updated successfully!')
            return redirect('coordinator:timetable_list')
        except Exception as e:
            messages.error(request, f'Error updating timetable: {str(e)}')
            return redirect('coordinator:timetable_edit', pk=pk)

    # GET — prefill the create form
    program_choices = [
        ('FSL', 'Future Skills Lab (FSL)'),
        ('CSL plus', 'CSL Plus'),
        ('CSL foundation', 'CSL Foundation'),
    ]
    existing_slots = [
        {
            'day': slot.day_of_week,
            'start': _fmt_time(slot.start_time),
            'end': _fmt_time(slot.end_time),
        }
        for slot in timetable.slots.all()
    ]
    context = {
        'editing': timetable,
        'editing_slots_json': json.dumps(existing_slots),
        'assigned_schools': assigned_schools,
        'thinking_coaches': thinking_coaches,
        'grade_choices': GRADE_CHOICES,
        'academic_year_choices': ACADEMIC_YEAR_CHOICES,
        'day_choices': DAY_CHOICES,
        'program_choices': program_choices,
    }
    return render(request, 'coordinator/timetable-upload.html', context)


@login_required
@user_passes_test(is_coordinator)
def timetable_delete(request, pk):
    """Delete a schedule (POST only). Scoped to coordinator's schools."""
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('coordinator:timetable_list')

    assigned_schools = _coordinator_schools(request)
    school_ids = assigned_schools.values_list('id', flat=True)
    timetable = Timetable.objects.filter(id=pk, school_id__in=school_ids).first()
    if not timetable:
        messages.error(request, 'Timetable not found or not in your assigned schools.')
        return redirect('coordinator:timetable_list')

    timetable.delete()
    messages.success(request, 'Timetable deleted successfully!')
    return redirect('coordinator:timetable_list')


@login_required
@user_passes_test(is_coordinator)
def timetable_upload(request):
    """Upload a schedule for a school class.
    Flow: select school -> assign thinking coach -> grade + division ->
    academic year -> program -> upload schedule file -> notes."""
    assigned_schools = _coordinator_schools(request)
    thinking_coaches = User.objects.filter(role='THINKING_COACH').order_by('first_name', 'username')

    if request.method == 'POST':
        school_id = request.POST.get('school', '').strip()
        coach_id = request.POST.get('thinking_coach', '').strip()
        grade = request.POST.get('grade', '').strip()
        division = request.POST.get('division', '').strip()
        academic_year = request.POST.get('academic_year', '').strip()
        program = request.POST.get('program', '').strip()
        start_date = request.POST.get('start_date', '').strip() or None
        end_date = request.POST.get('end_date', '').strip() or None
        notes = request.POST.get('notes', '').strip()
        schedule_file = request.FILES.get('schedule_file')

        # Validation
        if not school_id or not grade or not division:
            messages.error(request, 'School, grade and section are required.')
            return redirect('coordinator:timetable_upload')

        # Ensure the selected school belongs to this SRM
        school = assigned_schools.filter(id=school_id).first()
        if not school:
            messages.error(request, 'Invalid school selection.')
            return redirect('coordinator:timetable_upload')

        coach = None
        if coach_id:
            coach = thinking_coaches.filter(id=coach_id).first()

        # Auto-fill program from school's skill program if not provided
        if not program:
            program = school.get_skill_program_display() if school.skill_program else ''

        try:
            timetable = Timetable.objects.create(
                school=school,
                thinking_coach=coach,
                grade=grade,
                division=division,
                academic_year=academic_year or '2025-2026',
                program=program,
                start_date=start_date,
                end_date=end_date,
                schedule_file=schedule_file,
                notes=notes,
                created_by=request.user,
            )

            # Optional structured slot rows (additive nice-to-have)
            _save_slots(request, timetable)

            messages.success(request, 'Timetable uploaded successfully!')
            return redirect('coordinator:timetable_list')
        except Exception as e:
            messages.error(request, f'Error uploading timetable: {str(e)}')
            return redirect('coordinator:timetable_upload')

    program_choices = [
        ('FSL', 'Future Skills Lab (FSL)'),
        ('CSL plus', 'CSL Plus'),
        ('CSL foundation', 'CSL Foundation'),
    ]
    context = {
        'assigned_schools': assigned_schools,
        'thinking_coaches': thinking_coaches,
        'grade_choices': GRADE_CHOICES,
        'academic_year_choices': ACADEMIC_YEAR_CHOICES,
        'day_choices': DAY_CHOICES,
        'program_choices': program_choices,
    }
    return render(request, 'coordinator/timetable-upload.html', context)


# ============================================================
# ASSIGN THINKING COACHES (#12)
# ============================================================
@login_required
@user_passes_test(is_coordinator)
def assign_coaches(request):
    """List classes of the coordinator's assigned schools and assign a Thinking Coach to each.
    All operations are scoped to the coordinator's assigned schools for security."""
    assigned_schools = _coordinator_schools(request)
    school_ids = list(assigned_schools.values_list('id', flat=True))
    thinking_coaches = User.objects.filter(role='THINKING_COACH').order_by('first_name', 'username')

    if request.method == 'POST':
        class_id = request.POST.get('class_id', '').strip()
        coach_id = request.POST.get('thinking_coach', '').strip()

        # Only allow classes within the coordinator's assigned schools
        klass = Class.objects.filter(id=class_id, school_id__in=school_ids).first()
        if not klass:
            messages.error(request, 'Class not found or not in your assigned schools.')
            return redirect('coordinator:assign_coaches')

        coach = None
        if coach_id:
            coach = thinking_coaches.filter(id=coach_id).first()
            if not coach:
                messages.error(request, 'Invalid coach selection.')
                return redirect('coordinator:assign_coaches')

        klass.thinking_coach = coach
        klass.save()
        messages.success(request, 'Coach assigned successfully!')
        return redirect('coordinator:assign_coaches')

    classes = (
        Class.objects
        .filter(school_id__in=school_ids)
        .select_related('school', 'thinking_coach')
        .order_by('school__school_name', 'grade', 'division')
    )

    context = {
        'classes': classes,
        'thinking_coaches': thinking_coaches,
        'total_schools': assigned_schools.count(),
    }
    return render(request, 'coordinator/assign-coaches.html', context)


# ============================================================
# SCHOOL DETAILS (#11)
# ============================================================
@login_required
@user_passes_test(is_coordinator)
def school_detail(request, school_id):
    """Read-only detail page for a single school. Restricted to the coordinator's assigned schools."""
    assigned_schools = _coordinator_schools(request)
    school = assigned_schools.filter(id=school_id).first()
    if not school:
        messages.error(request, 'School not found or not in your assigned schools.')
        return redirect('coordinator:school_list')

    class_count = Class.objects.filter(school=school).count()
    student_count = Student.objects.filter(school=school).count()
    parent_count = Parent.objects.filter(students__school=school).distinct().count()

    context = {
        'school': school,
        'class_count': class_count,
        'student_count': student_count,
        'parent_count': parent_count,
    }
    return render(request, 'coordinator/school-details.html', context)


# ============================================================
# BULK UPLOAD — Students & Parents (client issue #10)
# ============================================================
# Reuses the superadmin bulk-import machinery WITHOUT modifying it.
# We import the parsing config + per-role processors from
# superadmin.bulk_import and run the same thin parse+process loop the
# (superadmin-decorated) bulk_import view runs, so behavior stays identical.

# Roles the coordinator is allowed to bulk-upload.
COORDINATOR_BULK_ROLES = ('student', 'parent')


@login_required
@user_passes_test(is_coordinator)
def bulk_upload_page(request):
    """Coordinator bulk-upload landing page — Student & Parent cards only."""
    return render(request, 'coordinator/bulk-upload.html')


@login_required
@user_passes_test(is_coordinator)
def download_sample_view(request, role):
    """Serve the same Excel sample template superadmin serves, scoped to
    the roles a coordinator may import (student/parent)."""
    from django.http import JsonResponse
    from superadmin.bulk_import import _generate_excel
    from django.http import HttpResponse

    if role not in COORDINATOR_BULK_ROLES:
        return JsonResponse({'error': 'Invalid role'}, status=400)

    wb = _generate_excel(role)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sample_{role}_import.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_coordinator)
def bulk_import_view(request, role):
    """Bulk-import students/parents. Mirrors superadmin.bulk_import.bulk_import
    exactly (same parse + process loop + JsonResponse shape) but is gated to
    coordinators and only student/parent roles."""
    import csv
    import io
    from django.http import JsonResponse
    from superadmin.bulk_import import EXCEL_CONFIG, ROLE_PROCESSORS, _get_display_name

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    if role not in COORDINATOR_BULK_ROLES:
        return JsonResponse({'error': 'Invalid role'}, status=400)

    csv_file = request.FILES.get('csv_file')
    if not csv_file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    file_name = csv_file.name.lower()
    is_excel = file_name.endswith('.xlsx')
    is_csv = file_name.endswith('.csv')

    if not is_excel and not is_csv:
        return JsonResponse({'error': 'Please upload a CSV or Excel (.xlsx) file'}, status=400)

    try:
        if is_excel:
            from openpyxl import load_workbook
            wb = load_workbook(csv_file, data_only=True)
            ws = wb.active
            # Row 2 has field names, data starts from row 3
            field_row = [str(cell.value or '').strip() for cell in ws[2]]
            rows = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                if all(v is None or str(v).strip() == '' for v in row):
                    continue
                row_dict = {}
                for idx, field in enumerate(field_row):
                    if field:
                        row_dict[field] = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ''
                rows.append(row_dict)
        else:
            decoded = csv_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
    except Exception as e:
        return JsonResponse({'error': f'Error reading file: {str(e)}'}, status=400)

    if not rows:
        return JsonResponse({'error': 'File is empty or has no data rows'}, status=400)

    # Validate headers — only REQUIRED columns must be present.
    expected = set(EXCEL_CONFIG[role]['required_fields'])
    actual = set(rows[0].keys())
    missing = expected - actual
    if missing:
        return JsonResponse({'error': f'Missing required columns: {", ".join(sorted(missing))}'}, status=400)

    results = []
    success_count = 0
    fail_count = 0
    processor = ROLE_PROCESSORS[role]

    for i, row in enumerate(rows):
        row = {k: (v.strip() if v else '') for k, v in row.items()}
        try:
            processor(row, request.user)
            success_count += 1
            results.append({'row': i + 1, 'name': _get_display_name(row, role), 'status': 'success'})
        except Exception as e:
            fail_count += 1
            results.append({'row': i + 1, 'name': _get_display_name(row, role), 'status': 'failed', 'reason': str(e)})

    return JsonResponse({
        'total': len(rows),
        'success': success_count,
        'failed': fail_count,
        'results': results,
    })


# ============================================================
# COMING SOON PLACEHOLDER (#17)
# ============================================================
@login_required
@user_passes_test(is_coordinator)
def coming_soon(request):
    """Simple placeholder page for features not yet built."""
    return render(request, 'coordinator/coming-soon.html')


@login_required
@user_passes_test(is_coordinator)
def coordinator_announcements(request):
    """Announcements targeted to this Program Coordinator (scoped to assigned schools)."""
    from competencies.announcements import announcements_for_user
    announcements = announcements_for_user(request.user)
    announcements.sort(key=lambda a: a.created_at, reverse=True)
    return render(request, 'coordinator/announcements.html', {'announcements': announcements})
