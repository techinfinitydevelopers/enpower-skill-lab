"""
Bulk CSV Import for all user roles.
Handles: sample CSV download, CSV parsing, user creation with error tracking.
"""
import csv
import io
import json
import secrets
import string
from datetime import datetime

from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.conf import settings
from django.db import transaction
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from schools.models import School

User = get_user_model()


def is_superadmin(user):
    return user.is_authenticated and user.role == "SUPER_ADMIN"


def generate_password(length=12):
    chars = string.ascii_letters + string.digits + '!@#$%'
    return ''.join(secrets.choice(chars) for _ in range(length))


def _opt(val):
    """Return None for empty strings, otherwise stripped value."""
    if val and val.strip():
        return val.strip()
    return None


def _opt_int(val, default=0):
    """Parse optional integer field."""
    if val and val.strip():
        try:
            return int(val.strip())
        except ValueError:
            return default
    return default


def _opt_bool(val, default=True):
    """Parse optional boolean field (yes/true/1 = True)."""
    if val and val.strip():
        return val.strip().lower() in ('yes', 'true', '1')
    return default


# ============================================================
# SAMPLE CSV DEFINITIONS — ALL columns per role
# ============================================================

SAMPLE_DATA = {
    'school_admin': {
        'headers': [
            'full_name', 'email', 'phone', 'gender', 'school_name',
            'date_of_birth', 'address', 'city', 'state', 'pincode',
        ],
        'rows': [
            ['Rahul Sharma', 'rahul@example.com', '9876543210', 'Male', 'Delhi Public School',
             '1990-05-15', '123 Main St', 'Mumbai', 'Maharashtra', '400001'],
            ['Priya Patel', 'priya@example.com', '9876543211', 'Female', 'Delhi Public School',
             '1992-08-20', '456 Park Ave', 'Delhi', 'Delhi', '110001'],
        ],
    },

    'teacher': {
        'headers': [
            # A. Basic Information
            'full_name', 'gender', 'date_of_birth', 'blood_group', 'nationality',
            'aadhar_number', 'pan_number',
            # B. Professional Details
            'designation', 'qualification', 'specialization', 'total_experience',
            'skill_training_experience', 'previous_organizations', 'certifications',
            'languages_known', 'grades_taught', 'training_style',
            # C. Contact Information
            'mobile_number', 'alternate_number', 'official_email', 'personal_email',
            # D. Address Details
            'current_address', 'permanent_address', 'city', 'state', 'pin_code',
            # E. Skill Lab Work Details
            'skill_lab_center', 'branch_location', 'batch_timings', 'weekly_timetable',
            'student_groups', 'modules_assigned', 'active_classes', 'total_students',
            'dashboard_role', 'joining_date', 'contract_end_date', 'employment_type',
            # F. Emergency Information
            'emergency_contact_name', 'emergency_relation', 'emergency_mobile',
            'emergency_secondary', 'health_notes',
            # G. Compliance & Documentation
            'id_proof_submitted', 'address_proof_submitted', 'police_verification',
            'contract_uploaded', 'pan_aadhar_linked', 'bank_details_submitted',
            # H. Bank Details
            'bank_name', 'branch_name', 'bank_account_number', 'ifsc_code',
            # I. Additional Optional Data
            'hobbies', 'strength_areas', 'improvement_areas', 'training_resources',
            'achievements',
            # School
            'school_name',
        ],
        'rows': [
            [
                # A. Basic
                'Ankit Verma', 'Male', '1988-03-10', 'B+', 'Indian',
                '123456789012', 'ABCDE1234F',
                # B. Professional
                'enpower-trainer', 'B.Ed', 'Science Education', '5 years',
                '3 years', '', '',
                'English, Hindi', '6,7,8', 'interactive',
                # C. Contact
                '9876543212', '', 'ankit@example.com', '',
                # D. Address
                '789 Oak Rd', '', 'Pune', 'Maharashtra', '411001',
                # E. Skill Lab
                '', '', '', '',
                '', '', '', '',
                '', '2024-04-01', '', 'full-time',
                # F. Emergency
                'Suresh Verma', 'parent', '9876543200',
                '', '',
                # G. Compliance
                '', '', '', '', '', '',
                # H. Bank
                'SBI', 'Kothrud', '12345678901234', 'SBIN0001234',
                # I. Additional
                '', '', '', '', '',
                # School
                'Delhi Public School',
            ],
        ],
    },

    'student': {
        'headers': [
            # A. Basic Information
            'first_name', 'middle_name', 'last_name', 'gender', 'date_of_birth',
            'nationality', 'mother_tongue', 'blood_group', 'aadhar_number',
            # B. Academic Details
            'school_name', 'school_branch', 'student_class', 'division', 'roll_number',
            'academic_year', 'gr_number', 'previous_school', 'stream', 'school_board',
            # C. Contact Details
            'student_mobile', 'school_email', 'personal_email', 'address',
            # D. Skill Lab Details
            'enrollment_date', 'skills_enrolled', 'current_skill_level', 'assigned_trainer',
            'batch_timing', 'learning_style', 'interests_aptitude', 'preferred_language',
            'practice_hours', 'certificates_earned', 'badges_earned',
            # E. Health & Safety
            'medical_conditions', 'allergies', 'emergency_instructions',
            'doctor_name', 'doctor_contact', 'physical_limitations',
            # F. Emergency Contact
            'emergency_name', 'emergency_relationship', 'emergency_mobile',
            'emergency_alt_mobile', 'emergency_address',
            # G. Family / Sibling Details
            'sibling_1_name', 'sibling_1_class_school', 'sibling_1_skill_lab_id',
            'sibling_2_name', 'sibling_2_class_school', 'sibling_2_skill_lab_id',
            'sibling_3_name', 'sibling_3_class_school', 'sibling_3_skill_lab_id',
            # H. Parent Linking
            'parent_email',
        ],
        'rows': [
            [
                # A. Basic
                'Aarav', 'Kumar', 'Rao', 'Male', '2010-06-15',
                'Indian', 'Hindi', 'O+', '123456789012',
                # B. Academic
                'Delhi Public School', '', '8', 'A', '101',
                '2024-2025', 'GR1234567890', '', '', 'CBSE',
                # C. Contact
                '9876543250', 'aarav@school.com', '', '123 Main St Mumbai',
                # D. Skill Lab
                '2024-04-01', '', '', '',
                '', '', '', '',
                '', '', '',
                # E. Health
                '', '', '',
                '', '', '',
                # F. Emergency
                'Rajesh Rao', 'father', '9876543213',
                '', '',
                # G. Siblings
                '', '', '',
                '', '', '',
                '', '', '',
                # H. Parent Linking
                'rajesh@example.com',
            ],
        ],
    },

    'parent': {
        'headers': [
            # A. Primary Parent / Guardian Details
            'full_name', 'relation_to_student', 'mobile_number', 'alternate_mobile',
            'email', 'occupation', 'organization', 'education_level', 'id_proof',
            # B. Student Assignment (student_names is for reference only, student_emails used for linking)
            'student_names', 'student_emails',
            # C. Secondary Parent / Guardian
            'secondary_full_name', 'secondary_relation', 'secondary_mobile',
            'secondary_email', 'secondary_occupation', 'preferred_contact',
            # C. Contact & Address
            'residential_address', 'landmark', 'city', 'state', 'pin_code',
            'permanent_address',
            # D. Communication Preferences
            'contact_method', 'preferred_language', 'dnd_timings',
            'whatsapp_consent', 'photo_consent',
            # E. Financial & Administrative
            'fee_category', 'payment_mode', 'billing_email', 'gst_number',
            # F. Emergency Contacts
            'emergency_name', 'emergency_relation', 'emergency_phone',
            'emergency_address',
            # G. Parent Involvement
            'meeting_availability', 'volunteer_interest', 'parent_skills',
        ],
        'rows': [
            [
                # A. Primary
                'Rajesh Rao', 'father', '9876543213', '',
                'rajesh@example.com', 'Engineer', 'TCS', 'graduate', '',
                # B. Student Assignment
                'Rahul Rao, Priya Rao', 'rahul@school.com,priya@school.com',
                # C. Secondary
                'Sunita Rao', 'mother', '9876543299',
                '', '', 'primary',
                # C. Address
                '123 Main St', '', 'Mumbai', 'Maharashtra', '400001',
                '',
                # D. Communication
                'whatsapp', 'hindi', '',
                'yes', 'yes',
                # E. Financial
                'regular', '', '', '',
                # F. Emergency
                'Amit Rao', 'uncle', '9876543214',
                '',
                # G. Involvement
                '', '', '',
            ],
        ],
    },

    'coordinator': {
        'headers': [
            # Basic Information
            'full_name', 'gender', 'date_of_birth', 'blood_group', 'nationality',
            'aadhar_number', 'pan_number',
            # Professional Details
            'designation', 'qualification', 'specialization', 'total_experience',
            'program_management_exp', 'education_exp', 'previous_organizations',
            'languages_known', 'certifications',
            # Contact Information
            'mobile_number', 'alternate_number', 'official_email', 'personal_email',
            # Address Details
            'current_address', 'permanent_address', 'city', 'state', 'pincode',
            # Compliance & Documentation
            'id_proof', 'address_proof', 'police_verification',
            'passport_photo_uploaded', 'contract_uploaded', 'pan_aadhar_linked', 'nda_signed',
            # Program & Work Assignment
            'program_assigned', 'zone_assigned', 'branch_region',
            'reporting_manager', 'login_role',
            'joining_date', 'employment_type', 'contract_start_date', 'contract_end_date',
            # Bank & Payroll Details
            'bank_name', 'branch_name', 'account_number', 'ifsc_code',
            # Additional Optional Data
            'strength_areas', 'hobbies', 'work_style', 'tools_comfortable',
            'achievements', 'career_aspirations',
        ],
        'rows': [
            [
                # Basic
                'Meera Joshi', 'Female', '1985-07-22', 'A+', 'Indian',
                '123456789012', 'ABCDE1234F',
                # Professional
                'Program Coordinator', 'MBA', 'Education', '10 years',
                '5 years', '3 years', '',
                'English, Hindi', '',
                # Contact
                '9876543215', '', 'meera@example.com', '',
                # Address
                '321 Elm St', '', 'Bangalore', 'Karnataka', '560001',
                # Compliance
                '', '', 'Pending',
                'No', 'No', 'No', 'No',
                # Program & Work
                'neoRISE', 'South Zone', '',
                'Ramesh Kumar', '',
                '2024-01-15', 'Full-time', '', '',
                # Bank
                'SBI', 'Koramangala', '12345678901234', 'SBIN0001234',
                # Additional
                '', '', '', '',
                '', '',
            ],
        ],
    },
}

ROLE_LABELS = {
    'school_admin': 'School Admin',
    'teacher': 'Teacher',
    'student': 'Student',
    'parent': 'Parent',
    'coordinator': 'Program Coordinator',
}


# ============================================================
# EXCEL CONFIG: HEADER MAPS, DROPDOWNS, REQUIRED FIELDS
# ============================================================

EXCEL_CONFIG = {
    'school_admin': {
        'sheet_title': 'School Admin Import',
        'header_map': {
            'full_name': 'Full Name',
            'email': 'Email Address',
            'phone': 'Phone Number',
            'gender': 'Gender',
            'school_name': 'School Name',
            'date_of_birth': 'Date of Birth',
            'address': 'Address',
            'city': 'City',
            'state': 'State',
            'pincode': 'PIN Code',
        },
        'dropdowns': {
            'gender': ['male', 'female', 'other', 'prefer-not-to-say'],
        },
        'required_fields': {
            'full_name', 'email', 'phone', 'gender', 'school_name',
        },
    },

    'teacher': {
        'sheet_title': 'Teacher Import',
        'header_map': {
            'full_name': 'Full Name (as per Aadhar)',
            'gender': 'Gender',
            'date_of_birth': 'Date of Birth',
            'blood_group': 'Blood Group',
            'nationality': 'Nationality',
            'aadhar_number': 'Aadhar Number',
            'pan_number': 'PAN Number',
            'designation': 'Designation',
            'qualification': 'Qualification',
            'specialization': 'Specialization',
            'total_experience': 'Total Experience',
            'skill_training_experience': 'Skill Training Experience',
            'previous_organizations': 'Previous Organizations',
            'certifications': 'Certifications',
            'languages_known': 'Languages Known',
            'grades_taught': 'Grades Taught',
            'training_style': 'Training Style',
            'mobile_number': 'Mobile Number',
            'alternate_number': 'Alternate Number',
            'official_email': 'Official Email ID',
            'personal_email': 'Personal Email',
            'current_address': 'Current Address',
            'permanent_address': 'Permanent Address',
            'city': 'City',
            'state': 'State',
            'pin_code': 'PIN Code',
            'skill_lab_center': 'Skill Lab Center',
            'branch_location': 'Branch Location',
            'batch_timings': 'Batch Timings',
            'weekly_timetable': 'Weekly Timetable',
            'student_groups': 'Student Groups',
            'modules_assigned': 'Modules Assigned',
            'active_classes': 'Active Classes',
            'total_students': 'Total Students',
            'dashboard_role': 'Dashboard Role',
            'joining_date': 'Joining Date',
            'contract_end_date': 'Contract End Date',
            'employment_type': 'Employment Type',
            'emergency_contact_name': 'Emergency Contact Name',
            'emergency_relation': 'Emergency Relation',
            'emergency_mobile': 'Emergency Mobile',
            'emergency_secondary': 'Emergency Secondary Mobile',
            'health_notes': 'Health Notes',
            'id_proof_submitted': 'ID Proof Submitted',
            'address_proof_submitted': 'Address Proof Submitted',
            'police_verification': 'Police Verification',
            'contract_uploaded': 'Contract Uploaded',
            'pan_aadhar_linked': 'PAN-Aadhar Linked',
            'bank_details_submitted': 'Bank Details Submitted',
            'bank_name': 'Bank Name',
            'branch_name': 'Branch Name',
            'bank_account_number': 'Bank Account Number',
            'ifsc_code': 'IFSC Code',
            'hobbies': 'Hobbies',
            'strength_areas': 'Strength Areas',
            'improvement_areas': 'Improvement Areas',
            'training_resources': 'Training Resources',
            'achievements': 'Achievements',
            'school_name': 'School Name',
        },
        'dropdowns': {
            'gender': ['male', 'female', 'other'],
            'blood_group': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'],
            'designation': ['enpower-trainer', 'school-teacher', 'head-teacher', 'assistant-teacher', 'principal', 'coordinator'],
            'training_style': ['interactive', 'conceptual', 'activity-based', 'mixed'],
            'employment_type': ['full-time', 'part-time', 'visiting', 'contract'],
            'dashboard_role': ['coach', 'senior-coach', 'coordinator', 'admin'],
            'emergency_relation': ['spouse', 'parent', 'sibling', 'child', 'friend', 'other'],
            'id_proof_submitted': ['yes', 'no', 'pending'],
            'address_proof_submitted': ['yes', 'no', 'pending'],
            'police_verification': ['yes', 'no', 'pending'],
            'contract_uploaded': ['yes', 'no', 'pending'],
            'pan_aadhar_linked': ['yes', 'no', 'pending'],
            'bank_details_submitted': ['yes', 'no', 'pending'],
        },
        'required_fields': {
            'full_name', 'gender', 'date_of_birth', 'nationality',
            'designation', 'qualification', 'total_experience',
            'mobile_number', 'official_email',
            'current_address', 'city', 'state', 'pin_code',
            'joining_date', 'employment_type',
            'emergency_contact_name', 'emergency_relation', 'emergency_mobile',
            'school_name',
        },
    },

    'student': {
        'sheet_title': 'Student Import',
        'header_map': {
            'first_name': 'First Name',
            'middle_name': 'Middle Name',
            'last_name': 'Last Name',
            'gender': 'Gender',
            'date_of_birth': 'Date of Birth',
            'nationality': 'Nationality',
            'mother_tongue': 'Mother Tongue',
            'blood_group': 'Blood Group',
            'aadhar_number': 'Aadhar Number',
            'school_name': 'School Name',
            'school_branch': 'School Branch',
            'student_class': 'Class',
            'division': 'Division',
            'roll_number': 'Roll Number',
            'academic_year': 'Academic Year',
            'gr_number': 'GR / Admission Number',
            'previous_school': 'Previous School',
            'stream': 'Stream',
            'school_board': 'School Board',
            'student_mobile': 'Student Mobile',
            'school_email': 'School Email ID',
            'personal_email': 'Personal Email',
            'address': 'Address',
            'enrollment_date': 'Enrollment Date',
            'skills_enrolled': 'Skills Enrolled',
            'current_skill_level': 'Current Skill Level',
            'assigned_trainer': 'Assigned Trainer',
            'batch_timing': 'Batch Timing',
            'learning_style': 'Learning Style',
            'interests_aptitude': 'Interests / Aptitude',
            'preferred_language': 'Preferred Language',
            'practice_hours': 'Practice Hours',
            'certificates_earned': 'Certificates Earned',
            'badges_earned': 'Badges Earned',
            'medical_conditions': 'Medical Conditions',
            'allergies': 'Allergies',
            'emergency_instructions': 'Emergency Instructions',
            'doctor_name': 'Doctor Name',
            'doctor_contact': 'Doctor Contact',
            'physical_limitations': 'Physical Limitations',
            'emergency_name': 'Emergency Contact Name',
            'emergency_relationship': 'Emergency Relationship',
            'emergency_mobile': 'Emergency Mobile',
            'emergency_alt_mobile': 'Emergency Alt Mobile',
            'emergency_address': 'Emergency Address',
            'sibling_1_name': 'Sibling 1 Name',
            'sibling_1_class_school': 'Sibling 1 Class/School',
            'sibling_1_skill_lab_id': 'Sibling 1 Skill Lab ID',
            'sibling_2_name': 'Sibling 2 Name',
            'sibling_2_class_school': 'Sibling 2 Class/School',
            'sibling_2_skill_lab_id': 'Sibling 2 Skill Lab ID',
            'sibling_3_name': 'Sibling 3 Name',
            'sibling_3_class_school': 'Sibling 3 Class/School',
            'sibling_3_skill_lab_id': 'Sibling 3 Skill Lab ID',
            'parent_email': 'Parent Email (for Linking)',
        },
        'dropdowns': {
            'gender': ['male', 'female', 'other'],
            'blood_group': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'],
            'stream': ['science', 'commerce', 'arts', 'na'],
            'school_board': ['CBSE', 'ICSE', 'SSC', 'IB', 'IGCSE', 'other'],
            'current_skill_level': ['beginner', 'intermediate', 'advanced'],
            'learning_style': ['visual', 'auditory', 'kinesthetic', 'mixed'],
            'preferred_language': ['english', 'hindi', 'marathi', 'other'],
            'emergency_relationship': ['father', 'mother', 'guardian', 'uncle', 'aunt', 'grandparent', 'sibling', 'other'],
        },
        'required_fields': {
            'first_name', 'last_name', 'gender', 'date_of_birth', 'nationality',
            'school_name',
            'student_class', 'division', 'roll_number', 'academic_year', 'gr_number', 'school_board',
            'school_email',
            'enrollment_date',
            'emergency_name', 'emergency_relationship', 'emergency_mobile',
            'parent_email',
        },
    },

    'parent': {
        'sheet_title': 'Parent Import',
        'header_map': {
            'full_name': 'Full Name',
            'relation_to_student': 'Relation to Student',
            'mobile_number': 'Mobile Number',
            'alternate_mobile': 'Alternate Mobile',
            'email': 'Email Address',
            'occupation': 'Occupation',
            'organization': 'Organization',
            'education_level': 'Education Level',
            'id_proof': 'ID Proof',
            'student_names': 'Student Names (Reference Only)',
            'student_emails': 'Student School Emails (for Linking)',
            'secondary_full_name': 'Secondary Guardian Name',
            'secondary_relation': 'Secondary Relation',
            'secondary_mobile': 'Secondary Mobile',
            'secondary_email': 'Secondary Email',
            'secondary_occupation': 'Secondary Occupation',
            'preferred_contact': 'Preferred Contact Person',
            'residential_address': 'Residential Address',
            'landmark': 'Landmark',
            'city': 'City',
            'state': 'State',
            'pin_code': 'PIN Code',
            'permanent_address': 'Permanent Address',
            'contact_method': 'Contact Method',
            'preferred_language': 'Preferred Language',
            'dnd_timings': 'DND Timings',
            'whatsapp_consent': 'WhatsApp Consent',
            'photo_consent': 'Photo Consent',
            'fee_category': 'Fee Category',
            'payment_mode': 'Payment Mode',
            'billing_email': 'Billing Email',
            'gst_number': 'GST Number',
            'emergency_name': 'Emergency Contact Name',
            'emergency_relation': 'Emergency Relation',
            'emergency_phone': 'Emergency Phone',
            'emergency_address': 'Emergency Address',
            'meeting_availability': 'Meeting Availability',
            'volunteer_interest': 'Volunteer Interest',
            'parent_skills': 'Parent Skills / Expertise',
        },
        'dropdowns': {
            'relation_to_student': ['father', 'mother', 'guardian'],
            'education_level': ['high-school', 'diploma', 'graduate', 'post-graduate', 'doctorate', 'other'],
            'secondary_relation': ['mother', 'father', 'guardian', 'grandparent', 'uncle', 'aunt'],
            'preferred_contact': ['primary', 'secondary', 'both'],
            'contact_method': ['call', 'whatsapp', 'sms', 'email'],
            'preferred_language': ['english', 'hindi', 'marathi', 'gujarati', 'tamil', 'telugu', 'kannada', 'other'],
            'whatsapp_consent': ['yes', 'no'],
            'photo_consent': ['yes', 'no'],
            'fee_category': ['regular', 'scholarship', 'concession'],
            'payment_mode': ['online', 'bank-transfer', 'cheque', 'cash', 'upi'],
            'emergency_relation': ['grandparent', 'uncle', 'aunt', 'sibling', 'neighbor', 'family-friend', 'other'],
            'meeting_availability': ['online', 'offline', 'both', 'not-available'],
            'volunteer_interest': ['yes', 'no'],
        },
        'required_fields': {
            'full_name', 'relation_to_student', 'mobile_number', 'email',
            'student_emails',
            'preferred_contact', 'residential_address', 'city', 'state', 'pin_code',
            'contact_method', 'preferred_language', 'fee_category',
            'emergency_name', 'emergency_relation', 'emergency_phone',
        },
    },

    'coordinator': {
        'sheet_title': 'Program Coordinator Import',
        'header_map': {
            'full_name': 'Full Name (as per Aadhar)',
            'gender': 'Gender',
            'date_of_birth': 'Date of Birth',
            'blood_group': 'Blood Group',
            'nationality': 'Nationality',
            'aadhar_number': 'Aadhar Number',
            'pan_number': 'PAN Number',
            'designation': 'Designation',
            'qualification': 'Qualification',
            'specialization': 'Specialization',
            'total_experience': 'Total Experience',
            'program_management_exp': 'Program Management Exp.',
            'education_exp': 'Education Experience',
            'previous_organizations': 'Previous Organizations',
            'languages_known': 'Languages Known',
            'certifications': 'Certifications',
            'mobile_number': 'Mobile Number',
            'alternate_number': 'Alternate Number',
            'official_email': 'Official Email ID',
            'personal_email': 'Personal Email',
            'current_address': 'Current Address',
            'permanent_address': 'Permanent Address',
            'city': 'City',
            'state': 'State',
            'pincode': 'PIN Code',
            'id_proof': 'ID Proof Submitted',
            'address_proof': 'Address Proof Submitted',
            'police_verification': 'Police Verification',
            'passport_photo_uploaded': 'Passport Photo Uploaded',
            'contract_uploaded': 'Contract Uploaded',
            'pan_aadhar_linked': 'PAN-Aadhar Linked',
            'nda_signed': 'NDA Signed',
            'program_assigned': 'Program Assigned',
            'zone_assigned': 'Zone Assigned',
            'branch_region': 'Branch / Region',
            'reporting_manager': 'Reporting Manager',
            'login_role': 'Login Role',
            'joining_date': 'Joining Date',
            'employment_type': 'Employment Type',
            'contract_start_date': 'Contract Start Date',
            'contract_end_date': 'Contract End Date',
            'bank_name': 'Bank Name',
            'branch_name': 'Branch Name',
            'account_number': 'Account Number',
            'ifsc_code': 'IFSC Code',
            'strength_areas': 'Strength Areas',
            'hobbies': 'Hobbies',
            'work_style': 'Work Style',
            'tools_comfortable': 'Tools Comfortable With',
            'achievements': 'Achievements',
            'career_aspirations': 'Career Aspirations',
        },
        'dropdowns': {
            'gender': ['male', 'female', 'other'],
            'blood_group': ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'],
            'designation': ['Program Coordinator', 'Project Coordinator'],
            'specialization': ['Operations', 'Education', 'Project Management', 'Others'],
            'employment_type': ['Full-time', 'Contract', 'Consultant'],
            'police_verification': ['Pending', 'In Progress', 'Completed'],
            'passport_photo_uploaded': ['Yes', 'No'],
            'contract_uploaded': ['Yes', 'No'],
            'pan_aadhar_linked': ['Yes', 'No'],
            'nda_signed': ['Yes', 'No'],
            'work_style': ['Field', 'Remote', 'Hybrid'],
        },
        'required_fields': {
            'full_name', 'gender', 'date_of_birth', 'nationality',
            'aadhar_number', 'pan_number',
            'designation', 'qualification', 'specialization', 'total_experience', 'languages_known',
            'mobile_number', 'official_email',
            'current_address', 'city', 'state', 'pincode',
            'id_proof',
            'program_assigned', 'joining_date', 'employment_type',
            'bank_name', 'branch_name', 'account_number', 'ifsc_code',
        },
    },
}


def _generate_excel(role):
    """Generate Excel (.xlsx) sample for any role with dropdowns and required highlighting."""
    config = EXCEL_CONFIG[role]
    data = SAMPLE_DATA[role]
    headers_raw = data['headers']
    header_map = config['header_map']
    dropdowns = config['dropdowns']
    required_fields = config['required_fields']
    headers_display = [header_map.get(h, h) for h in headers_raw]

    wb = Workbook()
    ws = wb.active
    ws.title = config['sheet_title']

    # -- Styles --
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    req_header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
    field_font = Font(name='Calibri', size=9, color='888888', italic=True)
    field_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
    req_field_font = Font(name='Calibri', size=9, color='B91C1C', italic=True)
    req_field_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
    data_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD'),
    )

    # Row 1: Human-readable headers (required → red bg + star)
    for col_idx, header in enumerate(headers_display, 1):
        field_name = headers_raw[col_idx - 1]
        is_req = field_name in required_fields
        display = f'{header} *' if is_req else header
        cell = ws.cell(row=1, column=col_idx, value=display)
        cell.font = header_font
        cell.fill = req_header_fill if is_req else header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Row 2: Field names (required → light red bg)
    for col_idx, field in enumerate(headers_raw, 1):
        is_req = field in required_fields
        cell = ws.cell(row=2, column=col_idx, value=field)
        cell.font = req_field_font if is_req else field_font
        cell.fill = req_field_fill if is_req else field_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Row 3+: Sample data
    for row_idx, row_data in enumerate(data['rows'], 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

    # -- Data Validations (dropdowns) for rows 3 to 1000 --
    for field_name, options in dropdowns.items():
        if field_name in headers_raw:
            col_idx = headers_raw.index(field_name) + 1
            col_letter = get_column_letter(col_idx)
            formula = '"' + ','.join(options) + '"'
            dv = DataValidation(type='list', formula1=formula, allow_blank=True)
            dv.error = f'Please select a valid option for {header_map.get(field_name, field_name)}'
            dv.errorTitle = 'Invalid Value'
            dv.prompt = f'Choose from: {", ".join(options)}'
            dv.promptTitle = header_map.get(field_name, field_name)
            dv.showInputMessage = True
            dv.showErrorMessage = True
            dv.sqref = f'{col_letter}3:{col_letter}1000'
            ws.add_data_validation(dv)

    # -- Column widths --
    for col_idx in range(1, len(headers_display) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(headers_display[col_idx - 1])), len(str(headers_raw[col_idx - 1])))
        for row_data in data['rows']:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)

    # Freeze top 2 rows
    ws.freeze_panes = 'A3'

    return wb


# ============================================================
# SAMPLE DOWNLOAD (CSV or Excel)
# ============================================================

@login_required
@user_passes_test(is_superadmin)
def download_sample_csv(request, role):
    if role not in SAMPLE_DATA:
        return JsonResponse({'error': 'Invalid role'}, status=400)

    # All roles → Excel with dropdowns, required highlighting
    wb = _generate_excel(role)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sample_{role}_import.xlsx"'
    wb.save(response)
    return response


# ============================================================
# BULK IMPORT PROCESSOR
# ============================================================

@login_required
@user_passes_test(is_superadmin)
def bulk_import(request, role):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    if role not in SAMPLE_DATA:
        return JsonResponse({'error': 'Invalid role'}, status=400)

    csv_file = request.FILES.get('csv_file')
    if not csv_file:
        return JsonResponse({'error': 'No file uploaded'}, status=400)

    file_name = csv_file.name.lower()
    is_excel = file_name.endswith('.xlsx')
    is_csv = file_name.endswith('.csv')

    if not is_excel and not is_csv:
        return JsonResponse({'error': 'Please upload a CSV or Excel (.xlsx) file'}, status=400)

    try:
        if is_excel:
            from openpyxl import load_workbook
            wb = load_workbook(csv_file, data_only=True)
            ws = wb.active
            # Row 2 has field names, data starts from row 3
            field_row = [str(cell.value or '').strip() for cell in ws[2]]
            rows = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                if all(v is None or str(v).strip() == '' for v in row):
                    continue
                row_dict = {}
                for idx, field in enumerate(field_row):
                    if field:
                        row_dict[field] = str(row[idx]).strip() if idx < len(row) and row[idx] is not None else ''
                rows.append(row_dict)
        else:
            decoded = csv_file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
    except Exception as e:
        return JsonResponse({'error': f'Error reading file: {str(e)}'}, status=400)

    if not rows:
        return JsonResponse({'error': 'File is empty or has no data rows'}, status=400)

    # Validate headers
    expected = set(SAMPLE_DATA[role]['headers'])
    actual = set(rows[0].keys())
    missing = expected - actual
    if missing:
        return JsonResponse({'error': f'Missing columns: {", ".join(sorted(missing))}'}, status=400)

    # Process each row
    results = []
    success_count = 0
    fail_count = 0

    processor = ROLE_PROCESSORS[role]

    for i, row in enumerate(rows):
        # Strip whitespace from all values
        row = {k: (v.strip() if v else '') for k, v in row.items()}
        try:
            processor(row, request.user)
            success_count += 1
            results.append({'row': i + 1, 'name': _get_display_name(row, role), 'status': 'success'})
        except Exception as e:
            fail_count += 1
            results.append({'row': i + 1, 'name': _get_display_name(row, role), 'status': 'failed', 'reason': str(e)})

    return JsonResponse({
        'total': len(rows),
        'success': success_count,
        'failed': fail_count,
        'results': results,
    })


def _get_display_name(row, role):
    if role == 'student':
        return f"{row.get('first_name', '')} {row.get('last_name', '')}".strip()
    return row.get('full_name', row.get('email', 'Row'))


# ============================================================
# PER-ROLE PROCESSORS
# ============================================================

def _process_school_admin(row, created_by):
    from school_admin.models import SchoolAdmin

    full_name = row.get('full_name', '')
    email = row.get('email', '')
    phone = row.get('phone', '')
    gender = row.get('gender', '')
    school_name = row.get('school_name', '')

    if not full_name:
        raise ValueError('full_name is required')
    if not email:
        raise ValueError('email is required')
    if not phone:
        raise ValueError('phone is required')
    if not gender:
        raise ValueError('gender is required')
    if not school_name:
        raise ValueError('school_name is required')

    school = School.objects.filter(school_name__iexact=school_name).first()
    if not school:
        raise ValueError(f'School "{school_name}" not found')

    if SchoolAdmin.objects.filter(school=school, is_active=True).exists():
        raise ValueError(f'School "{school_name}" already has an active admin')

    if User.objects.filter(username=email).exists():
        raise ValueError(f'Email "{email}" already exists')

    password = generate_password()
    name_parts = full_name.split(' ', 1)

    with transaction.atomic():
        user = User.objects.create_user(
            username=email, email=email, password=password,
            first_name=name_parts[0],
            last_name=name_parts[1] if len(name_parts) > 1 else '',
            role='SCHOOL_ADMIN',
        )

        SchoolAdmin.objects.create(
            user=user,
            full_name=full_name,
            email=email,
            phone=phone,
            gender=gender.lower(),
            school=school,
            date_of_birth=_parse_date(row.get('date_of_birth')),
            address=_opt(row.get('address')),
            city=_opt(row.get('city')),
            state=_opt(row.get('state')),
            pincode=_opt(row.get('pincode')),
            account_status='pending',
            is_active=True,
            temporary_password=password,
            created_by=created_by,
        )

    _send_welcome_email(email, full_name, password, 'School Admin')


def _process_teacher(row, created_by):
    from teacher.models import Teacher

    required = ['full_name', 'gender', 'date_of_birth', 'nationality',
                'designation', 'qualification', 'total_experience',
                'mobile_number', 'official_email',
                'current_address', 'city', 'state', 'pin_code',
                'joining_date', 'employment_type',
                'emergency_contact_name', 'emergency_relation', 'emergency_mobile',
                'school_name']

    for field in required:
        if not row.get(field):
            raise ValueError(f'{field} is required')

    email = row['official_email']
    if User.objects.filter(username=email).exists():
        raise ValueError(f'Email "{email}" already exists')

    school = None
    school_name = row.get('school_name', '')
    if school_name:
        school = School.objects.filter(school_name__iexact=school_name).first()
        if not school:
            raise ValueError(f'School "{school_name}" not found')

    password = generate_password()
    name_parts = row['full_name'].split(' ', 1)
    year = datetime.now().year
    emp_id = f"EMP{year}{''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))}"

    with transaction.atomic():
        user = User.objects.create_user(
            username=email, email=email, password=password,
            first_name=name_parts[0],
            last_name=name_parts[1] if len(name_parts) > 1 else '',
            role='THINKING_COACH',
        )

        Teacher.objects.create(
            user=user,
            school=school,
            employee_id=emp_id,
            # A. Basic Information
            full_name=row['full_name'],
            gender=row['gender'].lower(),
            date_of_birth=_parse_date(row['date_of_birth']),
            blood_group=_opt(row.get('blood_group')),
            nationality=row.get('nationality') or 'Indian',
            aadhar_number=_opt(row.get('aadhar_number')),
            pan_number=(_opt(row.get('pan_number')) or '').upper() or None,
            # B. Professional Details
            designation=row['designation'],
            qualification=row['qualification'],
            specialization=_opt(row.get('specialization')),
            total_experience=row['total_experience'],
            skill_training_experience=_opt(row.get('skill_training_experience')),
            previous_organizations=_opt(row.get('previous_organizations')),
            certifications=_opt(row.get('certifications')),
            languages_known=_opt(row.get('languages_known')),
            grades_taught=_opt(row.get('grades_taught')),
            training_style=_opt(row.get('training_style')),
            # C. Contact Information
            mobile_number=row['mobile_number'],
            alternate_number=_opt(row.get('alternate_number')),
            official_email=email,
            personal_email=_opt(row.get('personal_email')),
            # D. Address Details
            current_address=row['current_address'],
            permanent_address=_opt(row.get('permanent_address')),
            city=row['city'],
            state=row['state'],
            pin_code=row['pin_code'],
            # E. Skill Lab Work Details
            skill_lab_center=_opt(row.get('skill_lab_center')),
            branch_location=_opt(row.get('branch_location')),
            batch_timings=_opt(row.get('batch_timings')),
            weekly_timetable=_opt(row.get('weekly_timetable')),
            student_groups=_opt(row.get('student_groups')),
            modules_assigned=_opt(row.get('modules_assigned')),
            active_classes=_opt(row.get('active_classes')),
            total_students=_opt_int(row.get('total_students'), 0),
            dashboard_role=_opt(row.get('dashboard_role')),
            joining_date=_parse_date(row['joining_date']),
            contract_end_date=_parse_date(row.get('contract_end_date')),
            employment_type=row['employment_type'],
            # F. Emergency Information
            emergency_contact_name=row['emergency_contact_name'],
            emergency_relation=row['emergency_relation'],
            emergency_mobile=row['emergency_mobile'],
            emergency_secondary=_opt(row.get('emergency_secondary')),
            health_notes=_opt(row.get('health_notes')),
            # G. Compliance & Documentation
            id_proof_submitted=_opt(row.get('id_proof_submitted')),
            address_proof_submitted=_opt(row.get('address_proof_submitted')),
            police_verification=_opt(row.get('police_verification')),
            contract_uploaded=_opt(row.get('contract_uploaded')),
            pan_aadhar_linked=_opt(row.get('pan_aadhar_linked')),
            bank_details_submitted=_opt(row.get('bank_details_submitted')),
            # H. Bank Details
            bank_name=_opt(row.get('bank_name')),
            branch_name=_opt(row.get('branch_name')),
            bank_account_number=_opt(row.get('bank_account_number')),
            ifsc_code=(_opt(row.get('ifsc_code')) or '').upper() or None,
            # I. Additional Optional Data
            hobbies=_opt(row.get('hobbies')),
            strength_areas=_opt(row.get('strength_areas')),
            improvement_areas=_opt(row.get('improvement_areas')),
            training_resources=_opt(row.get('training_resources')),
            achievements=_opt(row.get('achievements')),
        )

    _send_welcome_email(email, row['full_name'], password, 'Thinking Coach')


def _process_student(row, created_by):
    from student.models import Student

    required = ['first_name', 'last_name', 'gender', 'date_of_birth', 'nationality',
                'school_name',
                'student_class', 'division', 'roll_number', 'academic_year', 'gr_number', 'school_board',
                'school_email',
                'enrollment_date',
                'emergency_name', 'emergency_relationship', 'emergency_mobile',
                'parent_email']

    for field in required:
        if not row.get(field):
            raise ValueError(f'{field} is required')

    email = row['school_email']
    gr_number = row['gr_number']

    if User.objects.filter(username=email).exists():
        raise ValueError(f'Email "{email}" already exists')

    if Student.objects.filter(gr_number=gr_number).exists():
        raise ValueError(f'GR Number "{gr_number}" already exists')

    school = None
    school_name = row.get('school_name', '')
    if school_name:
        school = School.objects.filter(school_name__iexact=school_name).first()
        if not school:
            raise ValueError(f'School "{school_name}" not found')

    password = generate_password()
    year = datetime.now().year
    reg_id = f"SKILL{year}{''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))}"

    with transaction.atomic():
        user = User.objects.create_user(
            username=email, email=email, password=password,
            first_name=row['first_name'],
            last_name=row['last_name'],
            role='STUDENT',
        )

        Student.objects.create(
            user=user,
            school=school,
            # A. Basic Information
            first_name=row['first_name'],
            middle_name=_opt(row.get('middle_name')),
            last_name=row['last_name'],
            gender=row['gender'].lower(),
            date_of_birth=_parse_date(row['date_of_birth']),
            nationality=row.get('nationality') or 'Indian',
            mother_tongue=_opt(row.get('mother_tongue')),
            blood_group=_opt(row.get('blood_group')),
            aadhar_number=_opt(row.get('aadhar_number')),
            # B. Academic Details
            school_name=school.school_name if school else (school_name or None),
            school_branch=_opt(row.get('school_branch')),
            student_class=row['student_class'],
            division=row['division'],
            roll_number=row['roll_number'],
            academic_year=row['academic_year'],
            gr_number=gr_number,
            previous_school=_opt(row.get('previous_school')),
            stream=_opt(row.get('stream')),
            school_board=row['school_board'],
            # C. Contact Details
            student_mobile=_opt(row.get('student_mobile')),
            school_email=email,
            personal_email=_opt(row.get('personal_email')),
            address=_opt(row.get('address')),
            # D. Skill Lab Details
            skill_lab_reg_id=reg_id,
            enrollment_date=_parse_date(row['enrollment_date']),
            skills_enrolled=_opt(row.get('skills_enrolled')),
            current_skill_level=_opt(row.get('current_skill_level')),
            assigned_trainer=_opt(row.get('assigned_trainer')),
            batch_timing=_opt(row.get('batch_timing')),
            learning_style=_opt(row.get('learning_style')),
            interests_aptitude=_opt(row.get('interests_aptitude')),
            preferred_language=_opt(row.get('preferred_language')),
            practice_hours=_opt_int(row.get('practice_hours'), 0),
            certificates_earned=_opt(row.get('certificates_earned')),
            badges_earned=_opt(row.get('badges_earned')),
            # E. Health & Safety
            medical_conditions=_opt(row.get('medical_conditions')),
            allergies=_opt(row.get('allergies')),
            emergency_instructions=_opt(row.get('emergency_instructions')),
            doctor_name=_opt(row.get('doctor_name')),
            doctor_contact=_opt(row.get('doctor_contact')),
            physical_limitations=_opt(row.get('physical_limitations')),
            # F. Emergency Contact
            emergency_name=row['emergency_name'],
            emergency_relationship=row['emergency_relationship'],
            emergency_mobile=row['emergency_mobile'],
            emergency_alt_mobile=_opt(row.get('emergency_alt_mobile')),
            emergency_address=_opt(row.get('emergency_address')),
            # G. Family / Sibling Details
            sibling_1_name=_opt(row.get('sibling_1_name')),
            sibling_1_class_school=_opt(row.get('sibling_1_class_school')),
            sibling_1_skill_lab_id=_opt(row.get('sibling_1_skill_lab_id')),
            sibling_2_name=_opt(row.get('sibling_2_name')),
            sibling_2_class_school=_opt(row.get('sibling_2_class_school')),
            sibling_2_skill_lab_id=_opt(row.get('sibling_2_skill_lab_id')),
            sibling_3_name=_opt(row.get('sibling_3_name')),
            sibling_3_class_school=_opt(row.get('sibling_3_class_school')),
            sibling_3_skill_lab_id=_opt(row.get('sibling_3_skill_lab_id')),
        )

        # H. Auto-link to parent if parent_email provided
        parent_email = _opt(row.get('parent_email'))
        if parent_email:
            from parent.models import Parent
            try:
                parent = Parent.objects.get(email=parent_email)
                parent.students.add(Student.objects.get(user=user))
            except Parent.DoesNotExist:
                pass  # Parent not imported yet — will auto-link when parent is imported later

    _send_welcome_email(email, f"{row['first_name']} {row['last_name']}", password, 'Student')


def _process_parent(row, created_by):
    from parent.models import Parent

    required = ['full_name', 'relation_to_student', 'mobile_number', 'email',
                'student_emails',
                'preferred_contact', 'residential_address', 'city', 'state', 'pin_code',
                'contact_method', 'preferred_language', 'fee_category',
                'emergency_name', 'emergency_relation', 'emergency_phone']

    for field in required:
        if not row.get(field):
            raise ValueError(f'{field} is required')

    email = row['email']
    if User.objects.filter(username=email).exists():
        raise ValueError(f'Email "{email}" already exists')

    password = generate_password()
    name_parts = row['full_name'].split(' ', 1)

    with transaction.atomic():
        user = User.objects.create_user(
            username=email, email=email, password=password,
            first_name=name_parts[0],
            last_name=name_parts[1] if len(name_parts) > 1 else '',
            role='PARENT',
        )

        parent = Parent.objects.create(
            user=user,
            # A. Primary Parent / Guardian Details
            full_name=row['full_name'],
            relation_to_student=row['relation_to_student'].lower(),
            mobile_number=row['mobile_number'],
            alternate_mobile=_opt(row.get('alternate_mobile')),
            email=email,
            occupation=_opt(row.get('occupation')),
            organization=_opt(row.get('organization')),
            education_level=_opt(row.get('education_level')),
            id_proof=_opt(row.get('id_proof')),
            # C. Secondary Parent / Guardian
            secondary_full_name=_opt(row.get('secondary_full_name')),
            secondary_relation=_opt(row.get('secondary_relation')),
            secondary_mobile=_opt(row.get('secondary_mobile')),
            secondary_email=_opt(row.get('secondary_email')),
            secondary_occupation=_opt(row.get('secondary_occupation')),
            preferred_contact=row.get('preferred_contact') or 'primary',
            # D. Contact & Address
            residential_address=row['residential_address'],
            landmark=_opt(row.get('landmark')),
            city=row['city'],
            state=row['state'],
            pin_code=row['pin_code'],
            permanent_address=_opt(row.get('permanent_address')),
            # E. Communication Preferences
            contact_method=row.get('contact_method') or 'whatsapp',
            preferred_language=row.get('preferred_language') or 'english',
            dnd_timings=_opt(row.get('dnd_timings')),
            whatsapp_consent=_opt_bool(row.get('whatsapp_consent'), True),
            photo_consent=_opt_bool(row.get('photo_consent'), True),
            # F. Financial & Administrative
            fee_category=row.get('fee_category') or 'regular',
            payment_mode=_opt(row.get('payment_mode')),
            billing_email=_opt(row.get('billing_email')),
            gst_number=_opt(row.get('gst_number')),
            # G. Emergency Contacts
            emergency_name=row['emergency_name'],
            emergency_relation=row['emergency_relation'],
            emergency_phone=row['emergency_phone'],
            emergency_address=_opt(row.get('emergency_address')),
            # H. Parent Involvement
            meeting_availability=_opt(row.get('meeting_availability')),
            volunteer_interest=_opt(row.get('volunteer_interest')),
            parent_skills=_opt(row.get('parent_skills')),
            # Status
            account_status='pending',
            is_active=True,
        )

        # B. Link students via school emails (student_names is reference only, ignored)
        student_emails_str = _opt(row.get('student_emails'))
        if student_emails_str:
            from student.models import Student
            emails = [e.strip() for e in student_emails_str.split(',') if e.strip()]
            for semail in emails:
                try:
                    student = Student.objects.get(school_email=semail)
                    parent.students.add(student)
                except Student.DoesNotExist:
                    pass  # Student not found — will auto-link when student is imported later

    _send_welcome_email(email, row['full_name'], password, 'Parent')


def _process_coordinator(row, created_by):
    from coordinator.models import ProgramCoordinator

    required = ['full_name', 'gender', 'date_of_birth', 'nationality',
                'aadhar_number', 'pan_number',
                'designation', 'qualification', 'specialization', 'total_experience', 'languages_known',
                'mobile_number', 'official_email',
                'current_address', 'city', 'state', 'pincode',
                'id_proof',
                'program_assigned', 'joining_date', 'employment_type',
                'bank_name', 'branch_name', 'account_number', 'ifsc_code']

    for field in required:
        if not row.get(field):
            raise ValueError(f'{field} is required')

    email = row['official_email']
    if User.objects.filter(username=email).exists():
        raise ValueError(f'Email "{email}" already exists')

    password = generate_password()
    name_parts = row['full_name'].split(' ', 1)
    year = datetime.now().year
    emp_id = f"EMP{year}{''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))}"

    with transaction.atomic():
        user = User.objects.create_user(
            username=email, email=email, password=password,
            first_name=name_parts[0],
            last_name=name_parts[1] if len(name_parts) > 1 else '',
            role='PROGRAM_COORDINATOR',
        )

        ProgramCoordinator.objects.create(
            user=user,
            employee_id=emp_id,
            # Basic Information
            full_name=row['full_name'],
            gender=row['gender'].lower(),
            date_of_birth=_parse_date(row['date_of_birth']),
            blood_group=_opt(row.get('blood_group')),
            nationality=row.get('nationality') or 'Indian',
            aadhar_number=row['aadhar_number'],
            pan_number=row['pan_number'].upper(),
            # Professional Details
            designation=row['designation'],
            qualification=row['qualification'],
            specialization=row['specialization'],
            total_experience=row['total_experience'],
            program_management_exp=_opt(row.get('program_management_exp')),
            education_exp=_opt(row.get('education_exp')),
            previous_organizations=_opt(row.get('previous_organizations')),
            languages_known=row['languages_known'],
            certifications=_opt(row.get('certifications')),
            # Contact Information
            mobile_number=row['mobile_number'],
            alternate_number=_opt(row.get('alternate_number')),
            official_email=email,
            personal_email=_opt(row.get('personal_email')),
            # Address Details
            current_address=row['current_address'],
            permanent_address=_opt(row.get('permanent_address')),
            city=row['city'],
            state=row['state'],
            pincode=row['pincode'],
            # Compliance & Documentation
            id_proof=_opt(row.get('id_proof')),
            address_proof=_opt(row.get('address_proof')),
            police_verification=row.get('police_verification') or 'Pending',
            passport_photo_uploaded=row.get('passport_photo_uploaded') or 'No',
            contract_uploaded=row.get('contract_uploaded') or 'No',
            pan_aadhar_linked=row.get('pan_aadhar_linked') or 'No',
            nda_signed=row.get('nda_signed') or 'No',
            # Program & Work Assignment
            program_assigned=_opt(row.get('program_assigned')),
            zone_assigned=_opt(row.get('zone_assigned')),
            branch_region=_opt(row.get('branch_region')),
            reporting_manager=_opt(row.get('reporting_manager')),
            login_role=_opt(row.get('login_role')),
            joining_date=_parse_date(row['joining_date']),
            employment_type=row['employment_type'],
            contract_start_date=_parse_date(row.get('contract_start_date')),
            contract_end_date=_parse_date(row.get('contract_end_date')),
            # Bank & Payroll Details
            bank_name=row['bank_name'],
            branch_name=row['branch_name'],
            account_number=row['account_number'],
            ifsc_code=row['ifsc_code'].upper(),
            # Additional Optional Data
            strength_areas=_opt(row.get('strength_areas')),
            hobbies=_opt(row.get('hobbies')),
            work_style=_opt(row.get('work_style')),
            tools_comfortable=_opt(row.get('tools_comfortable')),
            achievements=_opt(row.get('achievements')),
            career_aspirations=_opt(row.get('career_aspirations')),
        )

    _send_welcome_email(email, row['full_name'], password, 'Program Coordinator')


# ============================================================
# HELPERS
# ============================================================

def _parse_date(value):
    if not value or not value.strip():
        return None
    value = value.strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Invalid date format: "{value}". Use YYYY-MM-DD')


def _send_welcome_email(email, name, password, role_label):
    try:
        send_mail(
            subject=f'Welcome to Enpower Skill Lab — {role_label} Account',
            message=f'Hello {name},\n\nYour {role_label} account has been created.\n\nLogin: {email}\nPassword: {password}\n\nPlease change your password after first login.\n\nTeam Enpower Skill Lab',
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=True,
        )
    except Exception:
        pass  # Email failure should not block import


ROLE_PROCESSORS = {
    'school_admin': _process_school_admin,
    'teacher': _process_teacher,
    'student': _process_student,
    'parent': _process_parent,
    'coordinator': _process_coordinator,
}
