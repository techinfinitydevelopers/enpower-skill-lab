"""
School Admin Reports — the four grade-wise panels plus a full Excel download.

The panels are the same shared component the other roles use, scoped to this
principal's own school. The download writes the same figures to a workbook, one
sheet per panel, so what is on screen and what leaves as a file cannot disagree.

View-only throughout, per the presentation: this login belongs to the principal.
"""

from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.shortcuts import render

from competencies import report_panels
from .pages import _school_or_redirect
from .views import is_school_admin


def _panels_for(request, school):
    return report_panels.build([school.id], month=(request.GET.get('month') or None))


@login_required
@user_passes_test(is_school_admin)
def download_reports(request):
    """The panels on screen, with a button that exports the same figures."""
    school, bail = _school_or_redirect(request)
    if bail:
        return bail

    if request.GET.get('export') == 'xlsx':
        return _export_workbook(school, _panels_for(request, school))

    return render(request, 'school_admin/download-reports.html', {
        'school': school,
        'panels': _panels_for(request, school),
        'page_title': 'Download Reports',
    })


def _export_workbook(school, panels):
    """One sheet per panel, matching what the screen shows."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, color='FFFFFF', size=11)
    head_fill = PatternFill('solid', start_color='5B1F6F', end_color='5B1F6F')
    title_font = Font(bold=True, size=13)

    def sheet(title, subtitle, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]

        ws.cell(row=1, column=1, value=f'{school.school_name} — {subtitle}').font = title_font
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font, c.fill = head_font, head_fill
            c.alignment = Alignment(horizontal='center')
        for r_i, row in enumerate(rows, 4):
            for c_i, val in enumerate(row, 1):
                ws.cell(row=r_i, column=c_i, value=val)

        # Width from the longest cell, so nothing arrives as ####.
        for col in range(1, len(headers) + 1):
            longest = len(str(headers[col - 1]))
            for row in rows:
                if col - 1 < len(row):
                    longest = max(longest, len(str(row[col - 1])))
            ws.column_dimensions[get_column_letter(col)].width = min(longest + 4, 42)
        return ws

    d = panels['distribution']
    sheet('Student Distribution', 'Students by grade',
          ['Grade', 'Active students'],
          [[b['grade'], b['students']] for b in d['bars']] + [['Total', d['total']]],
          first=True)

    a = panels['attendance']
    month = panels.get('month_label') or 'all recorded months'
    sheet('Attendance', f'Attendance by grade — {month}',
          ['Grade', 'Attendance %'],
          [[b['grade'], b['percent']] for b in a['bars']]
          + ([['School average', a['school_average']]] if a['school_average'] is not None else []))

    c = panels['completion']
    sheet('Project Completion', 'Project completion by grade',
          ['Grade', 'Completion %'],
          [[b['grade'], b['percent']] for b in c['bars']]
          + ([['Overall', c['overall']]] if c['overall'] is not None else []))

    rows = []
    for row in (panels.get('profiles') or []):
        names = [p['name'] for p in row['profiles']] + [''] * 3
        counts = [p['students'] for p in row['profiles']] + [''] * 3
        rows.append([row['grade'], names[0], counts[0], names[1], counts[1], names[2], counts[2]])
    sheet('Top Skill Profiles', 'Top 3 skill profiles by grade',
          ['Grade', 'Profile 1', 'Students', 'Profile 2', 'Students', 'Profile 3', 'Students'],
          rows)

    filename = f'{school.school_name} reports.xlsx'.replace('/', '-')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
