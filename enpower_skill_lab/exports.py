"""
Excel export for the list screens that carry bulk buttons.

One registry, one view. Adding an export to a list means adding an entry here
rather than a new view per page, which is what kept these from existing: ten
pages had a Download Sample button and none had an export.

Each entry declares three things:

  roles    who may run it. A School Admin must not be able to export another
           school's parents, so the queryset is scoped from the request rather
           than trusted from a parameter.
  rows     a callable taking the request and returning the queryset, matching
           what that page shows -- including its own scoping and ordering.
  columns  (header, getter) pairs. The headers are the ones on screen, because
           that is what was asked for: whatever the list displays.

Getters are called with the object and must return something Excel can hold;
`_text` flattens the rest.
"""

from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.utils import timezone


# ── helpers ─────────────────────────────────────────────────────────────

def _text(value):
    """A cell value openpyxl will accept, for anything a getter returns."""
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def _active(obj):
    """Status as the lists show it, whichever field the model happens to use."""
    for attr in ('account_status', 'is_active', 'status'):
        if hasattr(obj, attr):
            value = getattr(obj, attr)
            if isinstance(value, bool):
                return 'Active' if value else 'Inactive'
            return str(value or '').title()
    return ''


def _school_name(obj):
    school = getattr(obj, 'school', None)
    return getattr(school, 'school_name', '') if school else ''


def _student_name(s):
    return ' '.join(p for p in (s.first_name, getattr(s, 'last_name', '')) if p)


def _student_class(s):
    cls = getattr(s, 'student_class', '') or ''
    div = getattr(s, 'division', '') or ''
    return f'{cls} - {div}' if cls and div else (cls or div)


def _children(parent):
    """The child names a parent row shows, joined."""
    return ', '.join(_student_name(s) for s in parent.students.all())


def _children_grades(parent):
    return ', '.join(str(getattr(s, 'student_class', '') or '')
                     for s in parent.students.all())


# ── scoping ─────────────────────────────────────────────────────────────

def _own_school(request):
    """The school this School Admin belongs to, or None."""
    profile = getattr(request.user, 'school_admin_profile', None)
    return getattr(profile, 'school', None) if profile else None


def _coach_school(request):
    profile = getattr(request.user, 'teacher_profile', None)
    return getattr(profile, 'school', None) if profile else None


def _coordinator_schools(request):
    from coordinator.views import _coordinator_schools as scoped
    return scoped(request)


# ── the registry ────────────────────────────────────────────────────────
# Keys are used in the URL, so they are stable strings rather than model names.

def _registry():
    """Built lazily so importing this module does not import every app."""
    from coordinator.models import ProgramCoordinator
    from parent.models import Parent
    from school_admin.models import SchoolAdmin
    from schools.models import School
    from student.models import Student
    from teacher.models import Teacher

    return {
        # ── Super Admin ─────────────────────────────────────────────────
        'schools': {
            'roles': ('SUPER_ADMIN',),
            'title': 'Schools',
            'rows': lambda r: School.objects.all().order_by('-created_at'),
            'columns': [
                ('School ID', lambda s: s.school_code),
                ('School Name', lambda s: s.school_name),
                ('Location', lambda s: ', '.join(p for p in (s.city, s.state) if p)),
                ('Contact Information', lambda s: s.school_email),
                ('Phone', lambda s: s.school_phone),
                ('Number of Students', lambda s: s.total_students or s.num_students),
                ('Status', _active),
            ],
        },
        'students': {
            'roles': ('SUPER_ADMIN',),
            'title': 'Students',
            'rows': lambda r: Student.objects.select_related('school').all()
                                     .order_by('-created_at'),
            'columns': [
                ('Student ID', lambda s: s.skill_lab_reg_id),
                ('Name', _student_name),
                ('Class', _student_class),
                ('School', _school_name),
                ('Status', _active),
            ],
        },
        'teachers': {
            'roles': ('SUPER_ADMIN',),
            'title': 'Thinking Coaches',
            'rows': lambda r: Teacher.objects.select_related('school').all()
                                     .order_by('-created_at'),
            'columns': [
                ('Teacher ID', lambda t: t.employee_id),
                ('Name', lambda t: t.full_name),
                ('Designation', lambda t: t.get_designation_display()
                    if hasattr(t, 'get_designation_display') else t.designation),
                ('Contact Information', lambda t: t.official_email),
                ('Mobile', lambda t: t.mobile_number),
                ('School', _school_name),
                ('Status', _active),
            ],
        },
        'parents': {
            'roles': ('SUPER_ADMIN',),
            'title': 'Parents',
            'rows': lambda r: Parent.objects.prefetch_related('students').all()
                                    .order_by('-created_at'),
            'columns': [
                ('Parent ID', lambda p: p.parent_id),
                ('Name', lambda p: p.full_name),
                ('Contact', lambda p: p.email),
                ('Mobile', lambda p: p.mobile_number),
                ("Child's Name", _children),
                ("Child's Grade", _children_grades),
                ('Status', _active),
            ],
        },
        'coordinators': {
            'roles': ('SUPER_ADMIN',),
            'title': 'Program Coordinators',
            'rows': lambda r: ProgramCoordinator.objects.all().order_by('-created_at'),
            'columns': [
                ('Employee ID', lambda c: c.employee_id),
                ('Name', lambda c: c.full_name),
                ('Designation', lambda c: c.designation),
                ('Contact', lambda c: c.official_email),
                ('Mobile', lambda c: c.mobile_number),
                ('Program Assigned', lambda c: c.program_assigned),
                ('Joining Date', lambda c: c.joining_date),
                ('Status', _active),
            ],
        },
        'school-admins': {
            'roles': ('SUPER_ADMIN',),
            'title': 'School Admins',
            'rows': lambda r: SchoolAdmin.objects.select_related('school').all()
                                         .order_by('-created_at'),
            'columns': [
                ('Admin ID', lambda a: a.id),
                ('Name', lambda a: a.full_name),
                ('School', _school_name),
                ('Contact Information', lambda a: a.email),
                ('Phone', lambda a: a.phone),
                ('Status', _active),
            ],
        },

        # ── Coordinator: only the schools mapped to them ────────────────
        'coordinator-schools': {
            'roles': ('PROGRAM_COORDINATOR',),
            'title': 'Schools',
            'rows': lambda r: _coordinator_schools(r).order_by('-created_at'),
            'columns': [
                ('School ID', lambda s: s.school_code),
                ('School Name', lambda s: s.school_name),
                ('Location', lambda s: ', '.join(p for p in (s.city, s.state) if p)),
                ('Contact Information', lambda s: s.school_email),
                ('Number of Students', lambda s: s.total_students or s.num_students),
                ('Status', _active),
            ],
        },

        # ── School Admin: only their own school ────────────────────────
        'my-students': {
            'roles': ('SCHOOL_ADMIN',),
            'title': 'Students',
            'rows': lambda r: Student.objects.filter(school=_own_school(r))
                                     .order_by('first_name', 'last_name')
            if _own_school(r) else Student.objects.none(),
            'columns': [
                ('Student ID', lambda s: s.skill_lab_reg_id),
                ('Name', _student_name),
                ('Class', _student_class),
                ('Email', lambda s: s.school_email),
                ('Status', _active),
            ],
        },
        'my-parents': {
            'roles': ('SCHOOL_ADMIN',),
            'title': 'Parents',
            'rows': lambda r: Parent.objects.filter(
                students__school=_own_school(r)).prefetch_related('students')
                .distinct().order_by('-created_at')
            if _own_school(r) else Parent.objects.none(),
            'columns': [
                ('Parent ID', lambda p: p.parent_id),
                ('Name', lambda p: p.full_name),
                ('Contact', lambda p: p.email),
                ("Child's Name", _children),
                ("Child's Grade", _children_grades),
                ('Status', _active),
            ],
        },

        # ── Coach: only their own school ───────────────────────────────
        'class-students': {
            'roles': ('THINKING_COACH',),
            'title': 'Students',
            'rows': lambda r: Student.objects.filter(
                school=_coach_school(r), is_active=True)
                .order_by('first_name', 'last_name')
            if _coach_school(r) else Student.objects.none(),
            'columns': [
                ('Student ID', lambda s: s.skill_lab_reg_id),
                ('Name', _student_name),
                ('Class', _student_class),
                ('Status', _active),
            ],
        },
    }


# ── the view ────────────────────────────────────────────────────────────

def export_list(request, key):
    """Stream one registry entry as an .xlsx."""
    registry = _registry()
    entry = registry.get(key)
    if entry is None:
        raise PermissionDenied('Unknown export')

    # Checked against the signed-in user, never against a parameter -- the key
    # names the list, it does not grant access to it.
    if getattr(request.user, 'role', None) not in entry['roles']:
        raise PermissionDenied('This export is not available to your role')

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = entry['title'][:31]

    headers = [h for h, _ in entry['columns']]
    ws.cell(row=1, column=1, value=entry['title']).font = Font(bold=True, size=13)
    ws.cell(row=1, column=max(2, len(headers)),
            value=f'Exported {timezone.localtime():%d %b %Y, %H:%M}').alignment = \
        Alignment(horizontal='right')

    head_font = Font(bold=True, color='FFFFFF', size=11)
    head_fill = PatternFill('solid', start_color='5B1F6F', end_color='5B1F6F')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font, cell.fill = head_font, head_fill
        cell.alignment = Alignment(horizontal='center')

    widths = [len(h) for h in headers]
    row_no = 4
    for obj in entry['rows'](request):
        for col, (_, getter) in enumerate(entry['columns'], 1):
            try:
                value = _text(getter(obj))
            except Exception:
                # One bad row must not lose the whole export.
                value = ''
            ws.cell(row=row_no, column=col, value=value)
            widths[col - 1] = max(widths[col - 1], len(str(value)))
        row_no += 1

    if row_no == 4:
        ws.cell(row=4, column=1, value='No records to export')

    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = min(width + 4, 46)
    ws.freeze_panes = 'A4'

    stamp = timezone.localtime().strftime('%Y-%m-%d')
    filename = f'{entry["title"]} {stamp}.xlsx'.replace('/', '-')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
